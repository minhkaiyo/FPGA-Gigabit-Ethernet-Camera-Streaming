import cv2
import numpy as np
import os
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from skimage.feature import local_binary_pattern, graycomatrix, graycoprops

# Link file.
LOG_IMG_DIR = r"D:\03. Thac si\04. AI\Image log"           # Thư mục chứa ảnh mẫu
LOGFILE_PATH = r"D:\03. Thac si\04. AI\Logfile4.xlsx"      # File lưu đặc trưng trung bình
TEST_DIR = r"D:\03. Thac si\04. AI\Test Pic"               # Thư mục test nhận diện

# HÀM TIỆN ÍCH
def imread_any(path: str) -> np.ndarray: 
    img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR) #Đọc cả ảnh tiếng Việt
    if img is None:
        img = cv2.imread(path, cv2.IMREAD_COLOR)
    return img

#Tách vật thể khỏi nền trắng
def segment_object(bgr: np.ndarray):
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (11, 9), 0) # làm mờ

    thresh = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 111, 5) # ngưỡng nhận diện và bù

    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 6))
    mask = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    mask_clean = np.zeros_like(mask)
    if contours:
        cv2.drawContours(mask_clean, [max(contours, key=cv2.contourArea)], -1, 255, -1)

    result = cv2.bitwise_and(bgr, bgr, mask=mask_clean)
    return mask_clean, result

# TRÍCH XUẤT ĐẶC TRƯNG
def features_from_masked(masked_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:

    gray = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2GRAY)

    # LBP gồm 26 đặc trưng
    P, R = 24, 3
    lbp = local_binary_pattern(gray, P=P, R=R, method="uniform")
    lbp_masked = lbp[mask > 0]
    n_bins = P + 2
    lbp_hist, _ = np.histogram(lbp_masked, bins=n_bins, range=(0, n_bins), density=True)

    # HSV histogram gồm 24 đặc trưng
    hsv = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)
    h, s, v = h[mask > 0], s[mask > 0], v[mask > 0]
    hist_h = np.histogram(h, bins=8, range=(0, 180), density=True)[0]
    hist_s = np.histogram(s, bins=8, range=(0, 256), density=True)[0]
    hist_v = np.histogram(v, bins=8, range=(0, 256), density=True)[0]
    hsv_hist = np.concatenate([hist_h, hist_s, hist_v])

    # Hu Moments gồm 7 đặc trưng
    moments = cv2.moments(mask)
    hu = cv2.HuMoments(moments).flatten()
    hu = -np.sign(hu) * np.log10(np.abs(hu) + 1e-12)  # Chuẩn hóa log scale để ổn định

    # GLMC (Gray-Level Co-occurrence Matrix) gồm 5 đặc trưng
    gray_norm = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    glcm = graycomatrix(gray_norm, distances=[1], angles=[0, np.pi/4, np.pi/2, 3*np.pi/4],
                        levels=256, symmetric=True, normed=True)
    contrast = graycoprops(glcm, 'contrast').mean()
    dissimilarity = graycoprops(glcm, 'dissimilarity').mean()
    homogeneity = graycoprops(glcm, 'homogeneity').mean()
    energy = graycoprops(glcm, 'energy').mean()
    correlation = graycoprops(glcm, 'correlation').mean()
    glcm_features = np.array([contrast, dissimilarity, homogeneity, energy, correlation])

    # Gộp tất cả đặc trưng
    feats = np.concatenate([lbp_hist, hsv_hist, hu, glcm_features]).astype(np.float32)
    return feats

def extract_features(img):
    """Tách nền và trích xuất toàn bộ đặc trưng."""
    mask, obj = segment_object(img)
    return features_from_masked(obj, mask)

# ========================= GIAO DIỆN CHÍNH =========================
root = Tk()
root.title("Nhận diện hoa quả phương pháp truyền thống")
root.geometry("1000x700")

lbl_status = Label(root, text="Sẵn sàng", font=("Arial", 13))
lbl_status.pack(pady=8)

# ========================= HÀM CHỨC NĂNG =========================
def reset_log():
    """Tạo lại file Logfile.xlsx từ thư mục Image log."""
    if not os.path.exists(LOG_IMG_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục: {LOG_IMG_DIR}")
        return
    lbl_status.config(text="Đang tạo Logfile.xlsx ...")
    root.update()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Tên sản phẩm"] + [f"f{i+1}" for i in range(76)])  # 7 Hu +26 LBP + 24 HSV + 7 Hu + 5 GLMC = 62 (cập nhật tự động nếu đổi số)

    for prod_name in os.listdir(LOG_IMG_DIR):
        prod_dir = os.path.join(LOG_IMG_DIR, prod_name)
        if not os.path.isdir(prod_dir):
            continue
        feats_list = []
        for img_name in os.listdir(prod_dir):
            img_path = os.path.join(prod_dir, img_name)
            img = imread_any(img_path)
            if img is None:
                continue
            feats = extract_features(img)
            feats_list.append(feats)
        if feats_list:
            feats_avg = np.mean(feats_list, axis=0)
            sheet.append([prod_name] + list(feats_avg))

    workbook.save(LOGFILE_PATH)
    lbl_status.config(text=f"✅ Đã tạo lại Logfile.xlsx tại {LOGFILE_PATH}")
    messagebox.showinfo("Hoàn tất", f"Đã cập nhật Logfile.xlsx ({LOGFILE_PATH})")

def identify_object_from_log(img):
    """Nhận diện vật thể bằng cách so sánh đặc trưng."""
    if not os.path.exists(LOGFILE_PATH):
        return "Chưa có Logfile.xlsx"
    feats = extract_features(img)
    wb = openpyxl.load_workbook(LOGFILE_PATH)
    ws = wb.active
    min_dist, best_name = float('inf'), "Không xác định"

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        vals = np.array(row[1:], dtype=float)
        dist = np.linalg.norm(vals - feats)
        if dist < min_dist:
            min_dist, best_name = dist, name
    return best_name

def identify_from_test_folder():
    """Nhận diện toàn bộ ảnh trong thư mục Test."""
    if not os.path.exists(TEST_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục Test: {TEST_DIR}")
        return
 #   if not os.path.exists(LOGFILE_PATH):
  #      messagebox.showerror("Lỗi", "Chưa có Logfile.xlsx để so sánh!")
   #     return

    results = []
    for file in os.listdir(TEST_DIR):
        if not file.lower().endswith((".jpg", ".png", ".jpeg")):
            continue
        img_path = os.path.join(TEST_DIR, file)
        img = imread_any(img_path)
        if img is None:
            continue
        name = identify_object_from_log(img)
        results.append((file, name))

    if not results:
        messagebox.showinfo("Không có ảnh Test file")
        return

    result_text = "\n".join([f"{f} → {n}" for f, n in results])
    lbl_status.config(text="Đã nhận diện xong ảnh trong thư mục Test")
    messagebox.showinfo("Kết quả nhận diện", result_text)

# ========================= NÚT GIAO DIỆN =========================
frm_buttons = Frame(root)
frm_buttons.pack(pady=30)

Button(frm_buttons, text="Reset Log", command=reset_log,
       bg="#E91E63", fg="white", font=("Arial", 14, "bold"), padx=20, pady=8).grid(row=0, column=0, padx=15)

Button(frm_buttons, text="Nhận diện ảnh trong thư mục Test", command=identify_from_test_folder,
       bg="#4CAF50", fg="white", font=("Arial", 14, "bold"), padx=20, pady=8).grid(row=0, column=1, padx=15)

# ========================= CHẠY GIAO DIỆN =========================
root.mainloop()