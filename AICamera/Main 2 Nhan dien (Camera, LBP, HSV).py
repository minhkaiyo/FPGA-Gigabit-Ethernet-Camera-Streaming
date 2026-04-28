import cv2
import numpy as np
import os, datetime
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from skimage.feature import local_binary_pattern

# ========================= CẤU HÌNH ĐƯỜNG DẪN =========================
SAVE_DIR = r"D:\03. Thac si\04. AI\Luu anh chup"          # Nơi lưu ảnh chụp
LOG_IMG_DIR = r"D:\03. Thac si\04. AI\Image log"           # Thư mục chứa ảnh mẫu từng loại
LOGFILE_PATH = r"D:\03. Thac si\04. AI\Logfile3.xlsx"      # File lưu đặc trưng trung bình
TEST_DIR = r"D:\03. Thac si\04. AI\Test Pic"               # Thư mục test nhận diện
os.makedirs(SAVE_DIR, exist_ok=True)

# ========================= HÀM TIỆN ÍCH =========================
def imread_any(path: str) -> np.ndarray:
    """Đọc ảnh kể cả khi tên file có tiếng Việt."""
    img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        img = cv2.imread(path, cv2.IMREAD_COLOR)
    return img

# Xử lý ảnh.
def segment_object(bgr: np.ndarray):
    """Tách vật thể khỏi nền trắng."""
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)              # Chuyển sang ảnh xám
    blur = cv2.GaussianBlur(gray, (7, 7), 0)                  # Làm mờ giảm nhiễu
    # Dùng ngưỡng thích nghi để tách vật thể (nền trắng → 0, vật thể → 1)
    thresh = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 51, 5)
    # Làm kín viền vật thể
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 6))
    mask = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)
    # Giữ lại contour lớn nhất (vật thể chính)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    mask_clean = np.zeros_like(mask)
    if contours:
        cv2.drawContours(mask_clean, [max(contours, key=cv2.contourArea)], -1, 255, -1)
    # Giữ lại phần ảnh có vật thể
    result = cv2.bitwise_and(bgr, bgr, mask=mask_clean)
    return mask_clean, result

# Nhận diện
def features_from_masked(masked_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    """Trích xuất đặc trưng (LBP + HSV histogram)"""
    gray = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2GRAY)       # Chuyển sang ảnh xám

    # ---- LBP (Local Binary Pattern) ----
    P, R = 24, 3                                              # Tham số LBP
    lbp = local_binary_pattern(gray, P=P, R=R, method="uniform")
    lbp_masked = lbp[mask > 0]                                # Chỉ lấy phần vật thể
    n_bins = P + 2                                            # Số lượng bins trong histogram
    lbp_hist, _ = np.histogram(lbp_masked, bins=n_bins, range=(0, n_bins), density=True)

    # ---- HSV Histogram ----
    hsv = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)
    h, s, v = h[mask > 0], s[mask > 0], v[mask > 0]
    hist_h = np.histogram(h, bins=8, range=(0, 180), density=True)[0]
    hist_s = np.histogram(s, bins=8, range=(0, 256), density=True)[0]
    hist_v = np.histogram(v, bins=8, range=(0, 256), density=True)[0]
    hsv_hist = np.concatenate([hist_h, hist_s, hist_v])

    # Gộp tất cả đặc trưng
    return np.concatenate([lbp_hist, hsv_hist]).astype(np.float32)

def extract_features(img):
    """Tách nền và trích xuất đặc trưng."""
    mask, obj = segment_object(img)
    return features_from_masked(obj, mask)

# ========================= GIAO DIỆN CHÍNH =========================
root = Tk()
root.title("Nhận diện hoa quả")
root.geometry("1200x950")

cap = cv2.VideoCapture(0)       # Mở camera
frame_data = None
lbl_video = Label(root)
lbl_video.pack()

lbl_status = Label(root, text="Sẵn sàng", font=("Arial", 13))
lbl_status.pack(pady=8)

# ========================= HÀM CHỨC NĂNG =========================
def reset_log():
    """Tạo lại file Logfile.xlsx từ thư mục Image log."""
    if not os.path.exists(LOG_IMG_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục: {LOG_IMG_DIR}")
        return
    lbl_status.config(text="🔄 Đang tạo Logfile.xlsx ...")
    root.update()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Tên sản phẩm"] + [f"f{i+1}" for i in range(56)])  # 56 = 26 (LBP) + 30 (HSV)

    for prod_name in os.listdir(LOG_IMG_DIR):
        prod_dir = os.path.join(LOG_IMG_DIR, prod_name)
        if not os.path.isdir(prod_dir):
            continue
        feats_list = []
        for img_name in os.listdir(prod_dir):
            img_path = os.path.join(prod_dir, img_name)
            img = imread_any(img_path)
            if img is None:
                continue
            feats = extract_features(img)
            feats_list.append(feats)
        if feats_list:
            feats_avg = np.mean(feats_list, axis=0)
            sheet.append([prod_name] + list(feats_avg))

    workbook.save(LOGFILE_PATH)
    lbl_status.config(text=f"Đã tạo lại Logfile.xlsx tại {LOGFILE_PATH}")
    messagebox.showinfo("Hoàn tất", f"Đã cập nhật Logfile.xlsx ({LOGFILE_PATH})")

def identify_object_from_log(img):
    """Nhận diện vật thể bằng cách so sánh đặc trưng."""
    if not os.path.exists(LOGFILE_PATH):
        return "Chưa có Logfile.xlsx"
    feats = extract_features(img)
    wb = openpyxl.load_workbook(LOGFILE_PATH)
    ws = wb.active
    min_dist, best_name = float('inf'), "Không xác định"

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        vals = np.array(row[1:], dtype=float)
        dist = np.linalg.norm(vals - feats)     # Tính khoảng cách Euclid
        if dist < min_dist:
            min_dist, best_name = dist, name
    return best_name

def capture_image():
    """Chụp ảnh và lưu lại."""
    global frame_data
    if frame_data is None:
        return
    choice = messagebox.askyesno("Lưu ảnh", "Bạn có muốn lưu ảnh này không?")
    if choice:
        dir_save = filedialog.askdirectory(initialdir=SAVE_DIR, title="Chọn thư mục lưu ảnh")
        if not dir_save:
            return
        filename = datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"
        filepath = os.path.join(dir_save, filename)
        cv2.imwrite(filepath, cv2.cvtColor(frame_data, cv2.COLOR_RGB2BGR))
        lbl_status.config(text=f"Đã lưu tại: {filepath}")
        messagebox.showinfo("Lưu ảnh", f"Ảnh đã lưu tại: {filepath}")
    else:
        lbl_status.config(text="Đã hủy lưu, quay lại camera")

def identify_from_test_folder():
    """Nhận diện tất cả ảnh trong thư mục Test."""
    if not os.path.exists(TEST_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục Test: {TEST_DIR}")
        return
    if not os.path.exists(LOGFILE_PATH):
        messagebox.showerror("Chưa có Logfile.xlsx để so sánh!")
        return

    results = []
    for file in os.listdir(TEST_DIR):
        if not file.lower().endswith((".jpg", ".png", ".jpeg")):
            continue
        img_path = os.path.join(TEST_DIR, file)
        img = imread_any(img_path)
        if img is None:
            continue
        name = identify_object_from_log(img)
        results.append((file, name))

    if not results:
        messagebox.showinfo("Kết quả", "Không tìm thấy ảnh nào trong thư mục Test.")
        return

    result_text = "\n".join([f"{f} → {n}" for f, n in results])
    lbl_status.config(text="✅ Đã nhận diện xong ảnh trong thư mục Test")
    messagebox.showinfo("Kết quả nhận diện", result_text)

def display_image(img_rgb):
    """Hiển thị ảnh camera lên giao diện."""
    img = Image.fromarray(img_rgb)
    imgtk = ImageTk.PhotoImage(image=img)
    lbl_video.imgtk = imgtk
    lbl_video.configure(image=imgtk)

def update_frame():
    """Cập nhật khung hình từ camera."""
    ret, frame = cap.read()
    if ret:
        rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        global frame_data
        frame_data = rgb
        display_image(rgb)
    lbl_video.after(30, update_frame)

def on_closing():
    """Khi đóng chương trình."""
    cap.release()
    root.destroy()

# ========================= NÚT GIAO DIỆN =========================
frm_buttons = Frame(root)
frm_buttons.pack(pady=15)

Button(frm_buttons, text="📸 Chụp ảnh", command=capture_image,
       bg="#4CAF50", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=0, padx=10)

Button(frm_buttons, text="🔄 Reset Log", command=reset_log,
       bg="#E91E63", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=1, padx=10)

Button(frm_buttons, text="🔍 Nhận diện", command=identify_from_test_folder,
       bg="#FF9800", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=2, padx=10)

# ========================= CHẠY CHƯƠNG TRÌNH =========================
root.protocol("WM_DELETE_WINDOW", on_closing)
update_frame()
root.mainloop()