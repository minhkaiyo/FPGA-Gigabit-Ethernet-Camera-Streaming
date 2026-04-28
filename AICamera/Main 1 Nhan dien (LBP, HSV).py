import cv2
import numpy as np
import os
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from skimage.feature import local_binary_pattern

# ========================= CẤU HÌNH ĐƯỜNG DẪN =========================
# Sử dụng đường dẫn tương đối với thư mục chứa script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_IMG_DIR = os.path.join(SCRIPT_DIR, "Image log")          # Thư mục chứa ảnh mẫu
LOGFILE_PATH = os.path.join(SCRIPT_DIR, "Logfile3.xlsx")     # File lưu đặc trưng trung bình
TEST_DIR = os.path.join(SCRIPT_DIR, "Test Pic")              # Thư mục test nhận diện

# ========================= HÀM TIỆN ÍCH =========================
def imread_any(path: str) -> np.ndarray:
    """Đọc ảnh kể cả khi tên file có tiếng Việt."""
    img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        img = cv2.imread(path, cv2.IMREAD_COLOR)
    return img

def segment_object(bgr: np.ndarray):
    """Tách vật thể khỏi nền trắng."""
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)              # Chuyển sang ảnh xám
    blur = cv2.GaussianBlur(gray, (11, 9), 0)                  # Làm mờ để giảm nhiễu

    # Ngưỡng thích nghi để tách nền trắng
    thresh = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 111, 5)

    # Làm kín vật thể
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 6))
    mask = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)

    # Lấy contour lớn nhất (vật thể chính)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    mask_clean = np.zeros_like(mask)
    if contours:
        cv2.drawContours(mask_clean, [max(contours, key=cv2.contourArea)], -1, 255, -1)

    # Giữ lại phần ảnh có vật thể
    result = cv2.bitwise_and(bgr, bgr, mask=mask_clean)
    return mask_clean, result

def features_from_masked(masked_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    """Trích xuất đặc trưng (LBP + HSV histogram).
    
    Returns:
        Vector đặc trưng 50 chiều: 26 (LBP) + 24 (HSV: 8H + 8S + 8V)
    """
    gray = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2GRAY)       # Chuyển sang ảnh xám

    # ---- LBP ----
    P, R = 24, 3                                              # Tham số LBP
    lbp = local_binary_pattern(gray, P=P, R=R, method="uniform")
    lbp_masked = lbp[mask > 0]                                # Chỉ lấy phần vật thể
    n_bins = P + 2                                            # 26 bins
    lbp_hist, _ = np.histogram(lbp_masked, bins=n_bins, range=(0, n_bins), density=True)

    # ---- HSV histogram ----
    hsv = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)
    h, s, v = h[mask > 0], s[mask > 0], v[mask > 0]
    hist_h = np.histogram(h, bins=8, range=(0, 180), density=True)[0]
    hist_s = np.histogram(s, bins=8, range=(0, 256), density=True)[0]
    hist_v = np.histogram(v, bins=8, range=(0, 256), density=True)[0]
    hsv_hist = np.concatenate([hist_h, hist_s, hist_v])       # 24 bins

    # Gộp tất cả đặc trưng lại (LBP 26 + HSV 24 = 50 đặc trưng)
    return np.concatenate([lbp_hist, hsv_hist]).astype(np.float32)

def extract_features(img):
    """Tách nền và trích xuất đặc trưng."""
    mask, obj = segment_object(img)
    return features_from_masked(obj, mask)

# ========================= GIAO DIỆN CHÍNH =========================
root = Tk()
root.title("Nhận diện hoa quả")
root.geometry("1000x700")

lbl_status = Label(root, text="Sẵn sàng", font=("Arial", 13))
lbl_status.pack(pady=8)

# ========================= HÀM CHỨC NĂNG =========================
def reset_log():
    """Tạo lại file Logfile.xlsx từ thư mục Image log."""
    if not os.path.exists(LOG_IMG_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục: {LOG_IMG_DIR}")
        return
    lbl_status.config(text="Đang tạo Logfile.xlsx ...")
    root.update()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 50 đặc trưng = 26 (LBP) + 24 (HSV: 8H + 8S + 8V)
    sheet.append(["Tên sản phẩm"] + [f"f{i+1}" for i in range(50)])

    for prod_name in os.listdir(LOG_IMG_DIR):
        prod_dir = os.path.join(LOG_IMG_DIR, prod_name)
        if not os.path.isdir(prod_dir):
            continue
        feats_list = []
        for img_name in os.listdir(prod_dir):
            img_path = os.path.join(prod_dir, img_name)
            img = imread_any(img_path)
            if img is None:
                continue
            feats = extract_features(img)
            feats_list.append(feats)
        if feats_list:
            feats_avg = np.mean(feats_list, axis=0)
            sheet.append([prod_name] + list(feats_avg))

    workbook.save(LOGFILE_PATH)
    lbl_status.config(text=f"✅ Đã tạo lại Logfile.xlsx tại {LOGFILE_PATH}")
    messagebox.showinfo("Hoàn tất", f"Đã cập nhật Logfile.xlsx ({LOGFILE_PATH})")

def identify_object_from_log(img):
    """Nhận diện vật thể bằng cách so sánh đặc trưng."""
    if not os.path.exists(LOGFILE_PATH):
        return "Chưa có Logfile.xlsx"
    feats = extract_features(img)
    wb = openpyxl.load_workbook(LOGFILE_PATH)
    ws = wb.active
    min_dist, best_name = float('inf'), "Không xác định"

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        vals = np.array(row[1:], dtype=float)
        dist = np.linalg.norm(vals - feats)
        if dist < min_dist:
            min_dist, best_name = dist, name
    return best_name

def identify_from_test_folder():
    """Nhận diện toàn bộ ảnh trong thư mục Test."""
    if not os.path.exists(TEST_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục Test: {TEST_DIR}")
        return
    if not os.path.exists(LOGFILE_PATH):
        messagebox.showerror("Lỗi", "Chưa có Logfile.xlsx để so sánh!")
        return

    results = []
    for file in os.listdir(TEST_DIR):
        if not file.lower().endswith((".jpg", ".png", ".jpeg")):
            continue
        img_path = os.path.join(TEST_DIR, file)
        img = imread_any(img_path)
        if img is None:
            continue
        name = identify_object_from_log(img)
        results.append((file, name))

    if not results:
        messagebox.showinfo("Kết quả", "Không tìm thấy ảnh nào trong thư mục Test.")
        return

    result_text = "\n".join([f"{f} → {n}" for f, n in results])
    lbl_status.config(text="✅ Đã nhận diện xong ảnh trong thư mục Test")
    # Hiển thị kết quả trong cửa sổ scrollable
    result_window = Toplevel(root)
    result_window.title("Kết quả nhận diện")
    result_window.geometry("600x400")
    
    scrollbar = Scrollbar(result_window)
    scrollbar.pack(side=RIGHT, fill=Y)
    
    text_widget = Text(result_window, yscrollcommand=scrollbar.set, font=("Consolas", 11))
    text_widget.pack(fill=BOTH, expand=True, padx=10, pady=10)
    text_widget.insert(END, result_text)
    text_widget.config(state=DISABLED)
    
    scrollbar.config(command=text_widget.yview)

def identify_single_image():
    """Nhận diện một ảnh được chọn."""
    if not os.path.exists(LOGFILE_PATH):
        messagebox.showerror("Lỗi", "Chưa có Logfile.xlsx để so sánh!\nHãy chạy 'Reset Log' trước.")
        return
    
    file_path = filedialog.askopenfilename(
        title="Chọn ảnh cần nhận diện",
        filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp"), ("All files", "*.*")]
    )
    
    if not file_path:
        return
    
    lbl_status.config(text="Đang nhận diện...")
    root.update()
    
    img = imread_any(file_path)
    if img is None:
        messagebox.showerror("Lỗi", "Không thể đọc ảnh!")
        lbl_status.config(text="Lỗi đọc ảnh")
        return
    
    try:
        result_name = identify_object_from_log(img)
        lbl_status.config(text=f"✅ Kết quả: {result_name}")
        messagebox.showinfo("Kết quả", f"Ảnh được nhận diện là:\n\n{result_name}", )
    except Exception as e:
        messagebox.showerror("Lỗi", f"Lỗi khi nhận diện:\n{str(e)}")
        lbl_status.config(text="Lỗi nhận diện")

# ========================= NÚT GIAO DIỆN =========================
frm_buttons = Frame(root)
frm_buttons.pack(pady=20)

Button(frm_buttons, text="🔄 Reset Log (Training)", command=reset_log,
       bg="#E91E63", fg="white", font=("Arial", 12, "bold"), padx=15, pady=10).grid(row=0, column=0, padx=10, pady=5)

Button(frm_buttons, text="🖼️ Nhận diện 1 ảnh", command=identify_single_image,
       bg="#2196F3", fg="white", font=("Arial", 12, "bold"), padx=15, pady=10).grid(row=0, column=1, padx=10, pady=5)

Button(frm_buttons, text="📁 Nhận diện thư mục Test", command=identify_from_test_folder,
       bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), padx=15, pady=10).grid(row=0, column=2, padx=10, pady=5)

# ========================= CHẠY GIAO DIỆN =========================
root.mainloop()
