import cv2
import numpy as np
import os, datetime
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from skimage.feature import local_binary_pattern, graycomatrix, graycoprops

# 1. File lưu (đường dẫn tương đối theo thư mục script)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SAVE_DIR = os.path.join(SCRIPT_DIR, "Luu_anh_chup")
LOG_IMG_DIR = os.path.join(SCRIPT_DIR, "Image_log")
LOGFILE_PATH = os.path.join(SCRIPT_DIR, "Logfile.xlsx")
TEST_DIR = os.path.join(SCRIPT_DIR, "Test_Pic")
os.makedirs(SAVE_DIR, exist_ok=True)
os.makedirs(LOG_IMG_DIR, exist_ok=True)
os.makedirs(TEST_DIR, exist_ok=True)

# 2. HÀM TIỆN ÍCH
def imread_any(path: str) -> np.ndarray: #Đọc ảnh hỗ trợ tên tiếng Việt
    img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR)
    if img is None:
        img = cv2.imread(path, cv2.IMREAD_COLOR)
    return img

def segment_object(bgr: np.ndarray):    # Tách vật thể khỏi nền
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (7,7), 0)
    thresh = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY_INV, 51, 5)
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7,6))
    mask = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2)
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    contours = sorted(contours, key=cv2.contourArea, reverse=True)
    mask_clean = np.zeros_like(mask)
    if len(contours) > 0:
        cv2.drawContours(mask_clean, [max(contours, key=cv2.contourArea)], -1, 255, -1)
    result = cv2.bitwise_and(bgr, bgr, mask=mask_clean)
    return mask_clean, result

# 3. TRÍCH XUẤT ĐẶC TRƯNG
def features_from_masked(masked_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2GRAY)

    # Hu Moments
    m = cv2.moments(mask, binaryImage=True)
    hu = cv2.HuMoments(m).flatten()
    hu = np.sign(hu) * np.log1p(np.abs(hu))

    # LBP
    P, R = 24, 3
    lbp = local_binary_pattern(gray, P=P, R=R, method="uniform")
    lbp_masked = lbp[mask > 0]
    n_bins = P + 2
    lbp_hist, _ = np.histogram(lbp_masked, bins=n_bins, range=(0, n_bins), density=True)

    # Haralick (GLCM)
    gray_m = gray.copy()
    gray_m[mask == 0] = 0
    quant = cv2.equalizeHist(gray_m)
    quant = (quant / 8).astype(np.uint8)
    distances = [1, 2, 3]
    angles = [0, np.pi/4, np.pi/2, 3*np.pi/4]
    glcm = graycomatrix(quant, distances=distances, angles=angles, levels=32, symmetric=True, normed=True)
    props = []
    for p in ["contrast", "dissimilarity", "homogeneity", "energy", "correlation", "ASM"]:
        vals = graycoprops(glcm, p)
        props.append(np.mean(vals))
    haralick = np.array(props, dtype=np.float32)

    # HSV histogram
    hsv = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2HSV)
    h, s, v = cv2.split(hsv)
    h = h[mask > 0]; s = s[mask > 0]; v = v[mask > 0]
    if h.size == 0:
        h, s, v = h.flatten(), s.flatten(), v.flatten()
    hist_h = np.histogram(h, bins=8, range=(0,180), density=True)[0]
    hist_s = np.histogram(s, bins=8, range=(0,256), density=True)[0]
    hist_v = np.histogram(v, bins=8, range=(0,256), density=True)[0]
    hsv_hist = np.concatenate([hist_h, hist_s, hist_v])

    return np.concatenate([hu, lbp_hist, haralick, hsv_hist]).astype(np.float32)

def extract_features(img):
    mask, obj = segment_object(img)
    return features_from_masked(obj, mask)

# 4. GIAO DIỆN CHÍNH
root = Tk()
root.title("Nhận diện hoa quả")
root.geometry("1200x950")

cap = cv2.VideoCapture(1)
frame_data = None
test_image = None
mode_segment = False

lbl_video = Label(root)
lbl_video.pack()

lbl_status = Label(root, text="🟢 Sẵn sàng", font=("Arial", 13))
lbl_status.pack(pady=8)

# ------------------ HÀM CHỨC NĂNG ------------------
def reset_log():
    if not os.path.exists(LOG_IMG_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục: {LOG_IMG_DIR}")
        return
    lbl_status.config(text="🔄 Đang tạo Logfile.xlsx ...")
    root.update()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Tên sản phẩm"] + [f"f{i+1}" for i in range(63)])

    for prod_name in os.listdir(LOG_IMG_DIR):
        prod_dir = os.path.join(LOG_IMG_DIR, prod_name)
        if not os.path.isdir(prod_dir):
            continue
        feats_list = []
        for img_name in os.listdir(prod_dir):
            img_path = os.path.join(prod_dir, img_name)
            img = imread_any(img_path)
            if img is None:
                continue
            feats = extract_features(img)
            feats_list.append(feats)
        if feats_list:
            feats_avg = np.mean(feats_list, axis=0)
            sheet.append([prod_name] + list(feats_avg))

    workbook.save(LOGFILE_PATH)
    lbl_status.config(text=f"Đã tạo lại Logfile.xlsx tại {LOGFILE_PATH}")
    messagebox.showinfo("Hoàn tất", f"Đã cập nhật Logfile.xlsx ({LOGFILE_PATH})")

def identify_object_from_log(img):
    if not os.path.exists(LOGFILE_PATH):
        return "Chưa có Logfile.xlsx"
    feats = extract_features(img)
    wb = openpyxl.load_workbook(LOGFILE_PATH)
    ws = wb.active
    min_dist, best_name = float('inf'), "Không xác định"

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        vals = np.array(row[1:], dtype=float)
        dist = np.linalg.norm(vals - feats)
        if dist < min_dist:
            min_dist, best_name = dist, name
    return best_name

def capture_image():
    """Chụp ảnh - popup có lưu không"""
    global frame_data
    if frame_data is None:
        return

    choice = messagebox.askyesno("Lưu ảnh", "Bạn có muốn lưu ảnh này không?")
    if choice:
        dir_save = filedialog.askdirectory(initialdir=SAVE_DIR, title="Chọn thư mục lưu ảnh")
        if not dir_save:
            return
        filename = datetime.datetime.now().strftime("%Y%m%d_%H%M%S") + ".jpg"
        filepath = os.path.join(dir_save, filename)
        cv2.imwrite(filepath, cv2.cvtColor(frame_data, cv2.COLOR_RGB2BGR))
        lbl_status.config(text=f"Đã lưu tại: {filepath}")
        messagebox.showinfo("Lưu ảnh", f"Ảnh đã lưu tại: {filepath}")
    else:
        lbl_status.config(text="Đã hủy lưu, quay lại camera")

def identify_from_test_folder():
    """Nhận diện ảnh trong thư mục Test"""
    if not os.path.exists(TEST_DIR):
        messagebox.showerror(f"Không tìm thấy thư mục Test: {TEST_DIR}")
        return
    if not os.path.exists(LOGFILE_PATH):
        messagebox.showerror("Chưa có Logfile.xlsx để so sánh!")
        return

    results = []
    for file in os.listdir(TEST_DIR):
        if not file.lower().endswith((".jpg", ".png", ".jpeg")):
            continue
        img_path = os.path.join(TEST_DIR, file)
        img = imread_any(img_path)
        if img is None:
            continue
        name = identify_object_from_log(img)
        results.append((file, name))

    if not results:
        messagebox.showinfo("Kết quả", "Không tìm thấy ảnh nào trong thư mục Test.")
        return

    result_text = "\n".join([f"{f} → {n}" for f, n in results])
    lbl_status.config(text="Đã nhận diện xong ảnh trong thư mục Test")
    messagebox.showinfo("Kết quả nhận diện", result_text)

def toggle_mode_segment():
    global mode_segment, test_image
    mode_segment = not mode_segment
    if mode_segment:
        file_path = filedialog.askopenfilename(title="Chọn ảnh cần test tách nền",
                                               filetypes=[("Ảnh", "*.jpg *.png *.jpeg")])
        if not file_path:
            mode_segment = False
            return
        img = imread_any(file_path)
        if img is None:
            messagebox.showerror("Lỗi", "Không đọc được ảnh!")
            mode_segment = False
            return
        test_image = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        display_image(test_image)
        lbl_status.config(text="Đang ở chế độ tách nền - Nhấn 'Tách nền'")
    else:
        lbl_status.config(text="Sẵn sàng")

def extract_background():
    global test_image
    if test_image is None:
        messagebox.showwarning("Hãy bật 'Mode tách nền' và chọn ảnh trước.")
        return
    bgr = cv2.cvtColor(test_image, cv2.COLOR_RGB2BGR)
    _, result = segment_object(bgr)
    result_rgb = cv2.cvtColor(result, cv2.COLOR_BGR2RGB)
    display_image(result_rgb)
    lbl_status.config(text="Tách nền thành công")

def display_image(img_rgb):
    img = Image.fromarray(img_rgb)
    imgtk = ImageTk.PhotoImage(image=img)
    lbl_video.imgtk = imgtk
    lbl_video.configure(image=imgtk)

def update_frame():
    if not mode_segment:
        ret, frame = cap.read()
        if ret:
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            global frame_data
            frame_data = rgb
            display_image(rgb)
    lbl_video.after(30, update_frame)

def on_closing():
    cap.release()
    root.destroy()

# ------------------ NÚT GIAO DIỆN ------------------
frm_buttons = Frame(root)
frm_buttons.pack(pady=15)

Button(frm_buttons, text="Chụp ảnh", command=capture_image,
       bg="#4CAF50", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=0, padx=10)
Button(frm_buttons, text="Reset Log", command=reset_log,
       bg="#E91E63", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=1, padx=10)
Button(frm_buttons, text="Mode tách nền", command=toggle_mode_segment,
       bg="#3F51B5", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=2, padx=10)
Button(frm_buttons, text="Tách nền", command=extract_background,
       bg="#009688", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=3, padx=10)
Button(frm_buttons, text="Nhận diện", command=identify_from_test_folder,
       bg="#FF9800", fg="white", font=("Arial", 14, "bold"), padx=15, pady=8).grid(row=0, column=4, padx=10)

# ------------------ CHẠY ------------------
root.protocol("WM_DELETE_WINDOW", on_closing)
update_frame()
root.mainloop()