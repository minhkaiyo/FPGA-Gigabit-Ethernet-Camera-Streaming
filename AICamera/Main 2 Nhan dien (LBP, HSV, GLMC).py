import cv2
import numpy as np
import os
from tkinter import *
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from skimage.feature import local_binary_pattern, graycomatrix, graycoprops

# ĐƯỜNG DẪN
LOG_IMG_DIR = r"D:\03. Thac si\04. AI\Image log"           # Thư mục chứa ảnh mẫu
LOGFILE_PATH = r"D:\03. Thac si\04. AI\Logfile5.xlsx"      # File lưu đặc trưng trung bình
TEST_DIR = r"D:\03. Thac si\04. AI\Test Pic"               # Thư mục test nhận diện

# HÀM TIỆN ÍCH
def imread_any(path: str) -> np.ndarray: 
    img = cv2.imdecode(np.fromfile(path, dtype=np.uint8), cv2.IMREAD_COLOR) #Đọc ảnh có tiếng Việt
    if img is None:
        img = cv2.imread(path, cv2.IMREAD_COLOR)
    return img

# Xử lý ảnh
def segment_object(bgr: np.ndarray):
    gray = cv2.cvtColor(bgr, cv2.COLOR_BGR2GRAY)
    blur = cv2.GaussianBlur(gray, (11, 9), 0) #0 là độ lệch chuẩn Gausian. Kenel càng lớn thì mất chi tiết nhiều do có nhiều đơn vị nhân vào.

    thresh = cv2.adaptiveThreshold(
        blur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, ## tính ngưỡng bằng trung bình có trọng số Gaussian
        cv2.THRESH_BINARY_INV, 111, 5) #111 là ngưỡng 111x111, lấy điểm trung tâm làm ngưỡng.

    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (7, 6))  # lọc nhiễu và khép kín viền
    mask = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel, iterations=2) #ảnh Thress Dilation → Erosion phóng to rồi thu -> xóa.

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)#tìm tất cả viền.RETR_EXTERNAL: Lấy ngoài cùng, CHAIN_APPROX_SIMPLE, lấy điểm quan trọng là góc và viền.
    mask_clean = np.zeros_like(mask)
    if contours:
        cv2.drawContours(mask_clean, [max(contours, key=cv2.contourArea)], -1, 255, -1) # số viền, màu viền, độ dày.

    result = cv2.bitwise_and(bgr, bgr, mask=mask_clean) #Áp mask lên ảnh gốc
    return mask_clean, result

# Trích xuất đặc trưng
def features_from_masked(masked_bgr: np.ndarray, mask: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2GRAY)

    # ---- LBP (Local Binary Pattern): 26 đặc trưng ----
    P, R = 24, 3 #24 là số điểm lân cận (5x5), 3 là bán kính.
    lbp = local_binary_pattern(gray, P=P, R=R, method="uniform") # thêm 2 bit Unitform và ununiform.
    lbp_masked = lbp[mask > 0]
    n_bins = P + 2
    lbp_hist, _ = np.histogram(lbp_masked, bins=n_bins, range=(0, n_bins), density=True)

    # HSV: 24 đặc trưng ----
    hsv = cv2.cvtColor(masked_bgr, cv2.COLOR_BGR2HSV) #chuyển sang H(Mau)S(Satuation-bao hoa)V(Value độ sáng.)
    h, s, v = cv2.split(hsv) # tách riêng biệt 3 kênh
    h, s, v = h[mask > 0], s[mask > 0], v[mask > 0]
    hist_h = np.histogram(h, bins=8, range=(0, 180), density=True)[0]
    hist_s = np.histogram(s, bins=8, range=(0, 256), density=True)[0]
    hist_v = np.histogram(v, bins=8, range=(0, 256), density=True)[0]
    hsv_hist = np.concatenate([hist_h, hist_s, hist_v]) #kết hợp vector

    # GLCM (Gray Level Co-occurrence Matrix): 5 đặc trưng
    gray_norm = cv2.normalize(gray, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8) #Chuan hóa ảnh về Gray 255
    glcm = graycomatrix(gray_norm, distances=[1],#tạo ma trận đồng xuất hiện xám.
                        angles=[0, np.pi/4, np.pi/2, 3*np.pi/4], # đi theo 4 hướng
                        levels=256, symmetric=True, normed=True)
    contrast = graycoprops(glcm, 'contrast').mean() # độ tương phản trung bình
    dissimilarity = graycoprops(glcm, 'dissimilarity').mean() #sự khác nhau giữa hai pixel
    homogeneity = graycoprops(glcm, 'homogeneity').mean() #độ đồng đều
    energy = graycoprops(glcm, 'energy').mean() #giá trị năng lượng
    correlation = graycoprops(glcm, 'correlation').mean() #Mối tương quan giữa các Pixel
    glcm_features = np.array([contrast, dissimilarity, homogeneity, energy, correlation])

    # Gộp tất cả đặc trưng
    feats = np.concatenate([lbp_hist, hsv_hist, glcm_features]).astype(np.float32)
    return feats

def extract_features(img):
    mask, obj = segment_object(img)
    return features_from_masked(obj, mask)

# GIAO DIỆN
root = Tk()
root.title("Nhận diện hoa quả:")
root.geometry("1000x700")

# CHỨC NĂNG
def reset_log(): #Tạo Logfile"""
    if not os.path.exists(LOG_IMG_DIR):
        messagebox.showerror(f"Không tìm thấy thư mục: {LOG_IMG_DIR}")
        return
    root.update()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Tên sản phẩm"] + [f"f{i+1}" for i in range(55)])  # 26 + 24 + 5 = 55

    for prod_name in os.listdir(LOG_IMG_DIR):
        prod_dir = os.path.join(LOG_IMG_DIR, prod_name)
        if not os.path.isdir(prod_dir):
            continue
        feats_list = []
        for img_name in os.listdir(prod_dir):
            img_path = os.path.join(prod_dir, img_name)
            img = imread_any(img_path)
            if img is None:
                continue
            feats = extract_features(img)
            feats_list.append(feats)
        if feats_list:
            feats_avg = np.mean(feats_list, axis=0)
            sheet.append([prod_name] + list(feats_avg))

    workbook.save(LOGFILE_PATH)
    messagebox.showinfo(f"Đã cập nhật Logfile.xlsx ({LOGFILE_PATH})")

def identify_object_from_log(img): #Nhận diện bằng cách so sánh đặc trưng và tính % giống."""
    if not os.path.exists(LOGFILE_PATH):
        return "Chưa có Logfile.xlsx", 0.0
    feats = extract_features(img)
    wb = openpyxl.load_workbook(LOGFILE_PATH)
    ws = wb.active

    distances = []
    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        vals = np.array(row[1:], dtype=float)
        dist = np.linalg.norm(vals - feats)
        names.append(name)
        distances.append(dist)

    if not distances:
        return "Không xác định", 0.0

    # Tìm khoảng cách nhỏ nhất
    min_dist = np.min(distances)
    best_name = names[np.argmin(distances)]

    # Chuẩn hóa ra phần trăm giống
    max_dist = np.max(distances)
    if max_dist == min_dist:
        similarity = 100.0
    else:
        similarity = 100 * (1 - (min_dist / max_dist))

    return best_name, similarity

def identify_from_test_folder(): #Nhận diện ảnh trong thư mục Test và hiển thị % giống
    if not os.path.exists(TEST_DIR):
        messagebox.showerror("Lỗi", f"Không tìm thấy thư mục Test: {TEST_DIR}")
        return

    results = []
    for file in os.listdir(TEST_DIR):
        if not file.lower().endswith((".jpg", ".png", ".jpeg")):
            continue
        img_path = os.path.join(TEST_DIR, file)
        img = imread_any(img_path)
        if img is None:
            continue
        name, sim = identify_object_from_log(img)
        results.append((file, name, sim))

    if not results:
        messagebox.showinfo("Không có ảnh Test file")
        return

    result_text = "\n".join([f"{f} → {n} ({s:.1f}%)" for f, n, s in results])
    messagebox.showinfo("Kết quả:", result_text)

# NÚT
frm_buttons = Frame(root)
frm_buttons.pack(pady=30)

Button(frm_buttons, text="Reset Log", command=reset_log,
       bg="#E91E63", fg="white", font=("Arial", 14, "bold"), padx=20, pady=8).grid(row=0, column=0, padx=15)

Button(frm_buttons, text="Nhận diện ảnh trong thư mục Test", command=identify_from_test_folder,
       bg="#4CAF50", fg="white", font=("Arial", 14, "bold"), padx=20, pady=8).grid(row=0, column=1, padx=15)

# CHẠY
root.mainloop()